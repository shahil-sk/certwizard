import os
import sys
import re
import json
import tempfile
import threading
import platform
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
except ImportError:
    messagebox.showerror("Missing Dependency", "Pillow is not installed.\nRun: pip install pillow")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    messagebox.showerror("Missing Dependency", "openpyxl is not installed.\nRun: pip install openpyxl")
    sys.exit(1)

try:
    from fpdf import FPDF
except ImportError:
    messagebox.showerror("Missing Dependency", "fpdf2 is not installed.\nRun: pip install fpdf2")
    sys.exit(1)


def resource_path(relative_path: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)


def px_to_mm(px: float) -> float:
    return px * 0.264583


# ---------------------------------------------------------------------------
# Theme
# ---------------------------------------------------------------------------
C = {
    "bg":       "#f5f6fa",
    "surface":  "#ffffff",
    "nav":      "#1a1d2e",
    "accent":   "#4f6ef7",
    "accent2":  "#3a56d4",
    "success":  "#2ecc71",
    "success2": "#27ae60",
    "border":   "#dfe1e8",
    "text":     "#1a1d2e",
    "subtext":  "#6b7280",
    "log_bg":   "#f8f9fb",
    "row_alt":  "#f0f2ff",
    "white":    "#ffffff",
}


def _btn(parent, text, command, bg, active_bg, font_size=9, bold=False, **kw):
    weight = "bold" if bold else "normal"
    return tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=C["white"], relief="flat", cursor="hand2",
        font=("Segoe UI", font_size, weight),
        activebackground=active_bg, activeforeground=C["white"],
        bd=0, highlightthickness=0, **kw
    )


def _label(parent, text, font_size=9, bold=False, color=None, **kw):
    weight = "bold" if bold else "normal"
    return tk.Label(
        parent, text=text,
        font=("Segoe UI", font_size, weight),
        fg=color or C["text"], bg=kw.pop("bg", C["surface"]),
        **kw
    )


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
class CertificateApp:

    APP_TITLE = "CertWizard"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(self.APP_TITLE)
        self.root.minsize(980, 640)
        self.root.configure(bg=C["bg"])

        self.original_image  = None
        self.display_image   = None
        self.scale_x = self.scale_y = 1.0
        self.template_path   = None
        self.excel_path      = None
        self.placeholders    = {}
        self._ph_images      = {}
        self._drag            = {}
        self.excel_data       = []
        self.fields           = []
        self.field_vars       = {}
        self.font_settings    = {}
        self.color_space      = tk.StringVar(value="RGB")
        self._gen_lock        = threading.Lock()

        self.available_fonts = self._load_fonts()
        self._setup_styles()
        self._build_ui()
        self._set_icon()

    # ------------------------------------------------------------------
    # Styles
    # ------------------------------------------------------------------
    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure(
            "Thin.Horizontal.TProgressbar",
            troughcolor=C["border"], background=C["accent"],
            thickness=6, borderwidth=0,
        )
        s.configure(
            "Flat.TCombobox",
            fieldbackground=C["surface"], background=C["surface"],
            foreground=C["text"], borderwidth=1, relief="flat",
        )
        s.map("Flat.TCombobox", fieldbackground=[("readonly", C["surface"])])
        s.configure(
            "Flat.Vertical.TScrollbar",
            background=C["border"], troughcolor=C["surface"],
            borderwidth=0, arrowsize=12,
        )

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        self._build_nav()
        self._build_statusbar()
        body = tk.Frame(self.root, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=12, pady=(0, 6))
        self._build_panel(body)
        self._build_canvas(body)

    def _build_nav(self):
        nav = tk.Frame(self.root, bg=C["nav"], height=48)
        nav.pack(fill="x")
        nav.pack_propagate(False)

        _label(nav, "CertWizard", font_size=13, bold=True,
               color="#ffffff", bg=C["nav"], padx=18).pack(side="left")

        _label(nav, "v2.1", font_size=8, color="#8892b0",
               bg=C["nav"]).pack(side="left", pady=(2, 0))

        right = tk.Frame(nav, bg=C["nav"])
        right.pack(side="right", padx=12)

        proj_btn = tk.Menubutton(
            right, text="Project", bg=C["nav"], fg="#c9d1d9",
            relief="flat", font=("Segoe UI", 9), padx=10,
            cursor="hand2", activebackground="#2d3250",
            activeforeground="white", bd=0,
        )
        proj_btn.menu = tk.Menu(
            proj_btn, tearoff=0,
            bg=C["surface"], fg=C["text"],
            activebackground=C["accent"], activeforeground="white",
            font=("Segoe UI", 9), bd=0, relief="flat",
        )
        proj_btn["menu"] = proj_btn.menu
        proj_btn.menu.add_command(label="Save Project", command=self.save_project)
        proj_btn.menu.add_command(label="Load Project", command=self.load_project)
        proj_btn.pack(side="left", padx=4)

        for text, cmd in (("Load Template", self.load_template),
                          ("Load Excel",    self.load_excel)):
            _btn(right, text, cmd, C["accent"], C["accent2"],
                 padx=12, pady=6).pack(side="left", padx=4)

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=C["nav"], height=24)
        bar.pack(side="bottom", fill="x")
        bar.pack_propagate(False)
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(
            bar, textvariable=self.status_var,
            bg=C["nav"], fg="#8892b0",
            font=("Segoe UI", 8), anchor="w", padx=14,
        ).pack(fill="x", expand=True)

    def _build_panel(self, parent):
        panel = tk.Frame(parent, width=300, bg=C["surface"],
                         highlightthickness=1,
                         highlightbackground=C["border"])
        panel.pack(side="left", fill="y", padx=(0, 10), pady=6)
        panel.pack_propagate(False)

        # Fields section
        hdr = tk.Frame(panel, bg=C["surface"], pady=10)
        hdr.pack(fill="x", padx=14)
        _label(hdr, "Certificate Fields", font_size=10, bold=True,
               bg=C["surface"]).pack(side="left")

        sep = tk.Frame(panel, bg=C["border"], height=1)
        sep.pack(fill="x", padx=14)

        self.fields_frame_outer = tk.Frame(panel, bg=C["surface"])
        self.fields_frame_outer.pack(fill="x", expand=False)

        self.toggle_frame = self.fields_frame_outer  # compat alias

        # Action buttons
        sep2 = tk.Frame(panel, bg=C["border"], height=1)
        sep2.pack(fill="x", padx=14, pady=(8, 0))

        btn_row = tk.Frame(panel, bg=C["surface"], pady=10)
        btn_row.pack(fill="x", padx=14)

        _btn(btn_row, "Preview",  self.preview_certificate,
             C["success"], C["success2"], font_size=9, bold=True,
             padx=14, pady=7).pack(side="left", fill="x", expand=True, padx=(0, 5))

        _btn(btn_row, "Generate", self.generate_certificates,
             C["accent"], C["accent2"], font_size=9, bold=True,
             padx=14, pady=7).pack(side="left", fill="x", expand=True)

        # Progress
        prog_frame = tk.Frame(panel, bg=C["surface"])
        prog_frame.pack(fill="x", padx=14, pady=(0, 8))
        self.progress = ttk.Progressbar(
            prog_frame, orient="horizontal", mode="determinate",
            style="Thin.Horizontal.TProgressbar",
        )
        self.progress.pack(fill="x")

        # Log
        sep3 = tk.Frame(panel, bg=C["border"], height=1)
        sep3.pack(fill="x", padx=14)

        log_hdr = tk.Frame(panel, bg=C["surface"], pady=8)
        log_hdr.pack(fill="x", padx=14)
        _label(log_hdr, "Generation Log", font_size=9, bold=True,
               color=C["subtext"], bg=C["surface"]).pack(side="left")

        log_wrap = tk.Frame(panel, bg=C["surface"])
        log_wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        self.info_text = tk.Text(
            log_wrap, height=8, wrap=tk.WORD,
            font=("Consolas", 8), bg=C["log_bg"], fg=C["subtext"],
            relief="flat", bd=0, padx=8, pady=6, state="disabled",
            highlightthickness=1, highlightbackground=C["border"],
        )
        self.info_text.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(log_wrap, orient="vertical",
                           command=self.info_text.yview,
                           style="Flat.Vertical.TScrollbar")
        sb.pack(side="right", fill="y")
        self.info_text.configure(yscrollcommand=sb.set)

    def _build_canvas(self, parent):
        wrap = tk.Frame(parent, bg=C["surface"],
                        highlightthickness=1,
                        highlightbackground=C["border"])
        wrap.pack(side="left", fill="both", expand=True, pady=6)
        self.canvas = tk.Canvas(wrap, bg=C["bg"], highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=1, pady=1)

    # ------------------------------------------------------------------
    # Icon
    # ------------------------------------------------------------------
    def _set_icon(self):
        for p in map(resource_path,
                     ("certgen.ico", "certgen.png", "icon.ico", "icon.png")):
            if not os.path.exists(p):
                continue
            try:
                if platform.system() == "Windows" and p.endswith(".ico"):
                    self.root.iconbitmap(p)
                else:
                    photo = ImageTk.PhotoImage(Image.open(p))
                    self.root.iconphoto(True, photo)
                    self._icon_ref = photo
                return
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Fonts
    # ------------------------------------------------------------------
    def _load_fonts(self) -> dict:
        fonts: dict = {}
        fonts_dir = resource_path("fonts")
        os.makedirs(fonts_dir, exist_ok=True)
        if os.path.isdir(fonts_dir):
            for f in sorted(os.listdir(fonts_dir)):
                if f.lower().endswith((".ttf", ".otf")):
                    fonts[os.path.splitext(f)[0]] = os.path.join(fonts_dir, f)
        if not fonts:
            fonts["Default"] = "arial.ttf"
        return dict(sorted(fonts.items(), key=lambda x: x[0].lower()))

    def _get_font(self, name: str, size: int) -> ImageFont.FreeTypeFont:
        try:
            return ImageFont.truetype(self.available_fonts.get(name, "arial.ttf"), size)
        except Exception:
            return ImageFont.load_default()

    # ------------------------------------------------------------------
    # Color helpers
    # ------------------------------------------------------------------
    def _hex_to_rgb(self, color) -> tuple:
        if isinstance(color, tk.StringVar):
            color = color.get()
        if color.startswith("cmyk("):
            c, m, y, k = map(float, color[5:-1].split(","))
            return (int(255*(1-c)*(1-k)), int(255*(1-m)*(1-k)), int(255*(1-y)*(1-k)))
        h = color.lstrip("#")
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    def _rgb_to_cmyk(self, hex_color: str) -> str:
        r, g, b = [x/255 for x in self._hex_to_rgb(hex_color)]
        k = 1 - max(r, g, b)
        if k == 1:
            return "cmyk(0.00,0.00,0.00,1.00)"
        c = (1-r-k)/(1-k); m = (1-g-k)/(1-k); y = (1-b-k)/(1-k)
        return f"cmyk({c:.2f},{m:.2f},{y:.2f},{k:.2f})"

    def _cmyk_to_hex(self, cmyk: str) -> str:
        c, m, y, k = map(float, cmyk[5:-1].split(","))
        return "#{:02x}{:02x}{:02x}".format(
            int(255*(1-c)*(1-k)), int(255*(1-m)*(1-k)), int(255*(1-y)*(1-k)))

    def choose_color(self, field: str):
        cur = self.font_settings[field]["color"].get()
        if self.color_space.get() == "RGB":
            init = cur if cur.startswith("#") else "#000000"
            chosen = colorchooser.askcolor(title=f"Pick color for {field}", initialcolor=init)
            if chosen[1]:
                self.font_settings[field]["color"].set(chosen[1])
        else:
            self._cmyk_picker(field, cur)
        self.update_preview(field)

    def _cmyk_picker(self, field: str, current: str):
        try:
            vals = list(map(float, current[5:-1].split(",")))
        except Exception:
            vals = [0.0, 0.0, 0.0, 0.0]

        win = tk.Toplevel(self.root)
        win.title(f"CMYK color for {field}")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        win.configure(bg=C["surface"])

        cvars = {ch: tk.DoubleVar(value=v)
                 for ch, v in zip("CMYK", vals)}
        preview = tk.Label(win, width=22, height=4, bg=C["surface"],
                           relief="flat", bd=0)
        preview.grid(row=4, column=0, columnspan=2, padx=16, pady=10)

        def refresh(*_):
            cmyk = "cmyk({:.2f},{:.2f},{:.2f},{:.2f})".format(
                cvars["C"].get(), cvars["M"].get(),
                cvars["Y"].get(), cvars["K"].get())
            self.font_settings[field]["color"].set(cmyk)
            preview.config(bg=self._cmyk_to_hex(cmyk))

        for row, (lbl, ch) in enumerate(
                zip(("Cyan", "Magenta", "Yellow", "Black"), "CMYK")):
            tk.Label(win, text=lbl, font=("Segoe UI", 9),
                     bg=C["surface"], fg=C["text"],
                     anchor="e").grid(row=row, column=0, padx=(16, 8),
                                      pady=5, sticky="e")
            tk.Scale(
                win, from_=0, to=1, resolution=0.01,
                variable=cvars[ch], command=refresh,
                orient="horizontal", length=200,
                bg=C["surface"], fg=C["text"],
                troughcolor=C["border"], highlightthickness=0,
                relief="flat",
            ).grid(row=row, column=1, padx=(0, 16))

        refresh()
        _btn(win, "Apply", win.destroy, C["accent"], C["accent2"],
             font_size=9, padx=24, pady=5).grid(
            row=5, column=0, columnspan=2, pady=(4, 14))
        self.root.wait_window(win)

    # ------------------------------------------------------------------
    # Load template / excel
    # ------------------------------------------------------------------
    def load_template(self, file_path=None):
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Select certificate template",
                filetypes=[("Images", "*.png *.jpg *.jpeg")])
        if not file_path:
            return
        try:
            self.template_path  = file_path
            self.original_image = Image.open(file_path).convert("RGBA")
            self._refresh_canvas()
            for f in self.fields:
                self.create_placeholder(f)
            self._status(f"Template: {os.path.basename(file_path)}")
        except Exception as exc:
            messagebox.showerror("Error", f"Cannot load template:\n{exc}")

    def _refresh_canvas(self):
        if self.original_image is None:
            return
        ow, oh = self.original_image.size
        mw, mh = 1000, 700
        ratio   = min(mw/ow, mh/oh)
        nw, nh  = int(ow*ratio), int(oh*ratio)
        self.scale_x, self.scale_y = ow/nw, oh/nh
        self.display_image = ImageTk.PhotoImage(
            self.original_image.resize((nw, nh), Image.LANCZOS))
        self.canvas.config(width=nw, height=nh)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")
        for field, data in list(self.placeholders.items()):
            self._draw_placeholder(field, data["x"], data["y"])

    def load_excel(self, file_path=None):
        if not file_path:
            file_path = filedialog.askopenfilename(
                title="Select Excel file",
                filetypes=[("Excel", "*.xlsx")])
        if not file_path:
            return
        try:
            wb    = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            header = [
                str(c.value).strip().lower()
                for c in next(sheet.iter_rows(min_row=1, max_row=1))
                if c.value is not None
            ]
            if not header:
                messagebox.showerror("Error", "No column headers found.")
                return

            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i >= len(header):
                        break
                    if isinstance(val, datetime):
                        val = val.strftime("%d-%m-%Y")
                    rec[header[i]] = str(val) if val is not None else ""
                if any(rec.values()):
                    data.append(rec)
            wb.close()

            if not data:
                messagebox.showwarning("Warning", "No data rows found.")
                return

            self.excel_path = file_path
            self.fields     = header
            self.excel_data = data

            default_font = next(iter(self.available_fonts))
            self.field_vars    = {f: tk.BooleanVar(value=True) for f in header}
            self.font_settings = {
                f: {
                    "size":      tk.IntVar(value=32),
                    "color":     tk.StringVar(value="#000000"),
                    "font_name": tk.StringVar(value=default_font),
                } for f in header
            }

            self._build_field_panel()
            if self.original_image:
                for f in self.fields:
                    self.create_placeholder(f)

            self._status(f"Excel loaded: {len(data)} records, {len(header)} fields")
        except Exception as exc:
            messagebox.showerror("Error", f"Cannot load Excel:\n{exc}")

    # ------------------------------------------------------------------
    # Field panel
    # ------------------------------------------------------------------
    def _build_field_panel(self):
        for w in self.fields_frame_outer.winfo_children():
            w.destroy()

        scroll_canvas = tk.Canvas(
            self.fields_frame_outer, height=260,
            bg=C["surface"], highlightthickness=0,
        )
        vsb = ttk.Scrollbar(
            self.fields_frame_outer, orient="vertical",
            command=scroll_canvas.yview,
            style="Flat.Vertical.TScrollbar",
        )
        inner = tk.Frame(scroll_canvas, bg=C["surface"])
        inner.bind(
            "<Configure>",
            lambda e: scroll_canvas.configure(
                scrollregion=scroll_canvas.bbox("all")),
        )
        scroll_canvas.create_window((0, 0), window=inner, anchor="nw")
        scroll_canvas.configure(yscrollcommand=vsb.set)

        for i, field in enumerate(self.fields):
            self._add_field_row(inner, field, alt=(i % 2 == 1))

        vsb.pack(side="right", fill="y")
        scroll_canvas.pack(side="left", fill="both", expand=True)

    def _add_field_row(self, parent, field: str, alt=False):
        bg = C["row_alt"] if alt else C["surface"]
        row = tk.Frame(parent, bg=bg, pady=8, padx=12)
        row.pack(fill="x")

        top = tk.Frame(row, bg=bg)
        top.pack(fill="x", pady=(0, 5))

        _label(top, field.title(), font_size=9, bold=True,
               bg=bg).pack(side="left")

        cb_vis = tk.Checkbutton(
            top, variable=self.field_vars[field],
            bg=bg, activebackground=bg,
            relief="flat", bd=0,
            command=lambda f=field: self.update_preview(f),
        )
        cb_vis.pack(side="right")
        _label(top, "Visible", font_size=8,
               color=C["subtext"], bg=bg).pack(side="right", padx=(0, 2))

        ctrl = tk.Frame(row, bg=bg)
        ctrl.pack(fill="x")

        _label(ctrl, "Size", font_size=8, color=C["subtext"], bg=bg).pack(side="left")
        spin = tk.Spinbox(
            ctrl, from_=8, to=300, width=4,
            textvariable=self.font_settings[field]["size"],
            font=("Segoe UI", 9), relief="flat", bd=1,
            command=lambda f=field: self.update_preview(f),
        )
        spin.bind("<Return>", lambda e, f=field: self.update_preview(f))
        spin.pack(side="left", padx=(3, 10))

        _label(ctrl, "Font", font_size=8, color=C["subtext"], bg=bg).pack(side="left")
        cb = ttk.Combobox(
            ctrl, values=list(self.available_fonts.keys()),
            textvariable=self.font_settings[field]["font_name"],
            width=12, state="readonly", style="Flat.TCombobox",
        )
        cb.bind("<<ComboboxSelected>>", lambda e, f=field: self.update_preview(f))
        cb.pack(side="left", padx=(3, 8))

        swatch = tk.Label(
            ctrl, width=2, height=1,
            bg=self.font_settings[field]["color"].get(),
            relief="flat", bd=1,
            highlightthickness=1, highlightbackground=C["border"],
        )
        swatch.pack(side="left", padx=(0, 3))
        self.font_settings[field]["_swatch"] = swatch
        _btn(
            ctrl, "Color",
            lambda f=field: self._pick_color(f),
            "#e8eaf0", C["border"],
            font_size=8, padx=7, pady=2,
        ).configure(fg=C["text"], activeforeground=C["text"])
        _btn(
            ctrl, "Color",
            lambda f=field: self._pick_color(f),
            "#e8eaf0", "#d0d4df",
            font_size=8, padx=7, pady=2,
        ).configure(fg=C["text"], activeforeground=C["text"])
        # replace double btn with single
        for w in ctrl.winfo_children():
            if isinstance(w, tk.Button) and w.cget("text") == "Color":
                w.destroy()
        color_btn = tk.Button(
            ctrl, text="Color",
            command=lambda f=field: self._pick_color(f),
            bg="#e8eaf0", fg=C["text"],
            relief="flat", bd=0, cursor="hand2",
            font=("Segoe UI", 8),
            activebackground="#d0d4df",
            padx=7, pady=2,
        )
        color_btn.pack(side="left")

    def _pick_color(self, field: str):
        self.choose_color(field)
        swatch = self.font_settings[field].get("_swatch")
        if swatch:
            col = self.font_settings[field]["color"].get()
            if col.startswith("cmyk("):
                col = self._cmyk_to_hex(col)
            try:
                swatch.config(bg=col)
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Placeholders
    # ------------------------------------------------------------------
    def _render_placeholder_image(self, field: str):
        s     = self.font_settings[field]
        size  = s["size"].get()
        color = s["color"].get()
        font  = self._get_font(s["font_name"].get(), size)
        text  = (self.excel_data[0].get(field, field) if self.excel_data else field) or field

        tmp  = Image.new("RGBA", (1, 1))
        draw = ImageDraw.Draw(tmp)
        tw   = max(int(draw.textlength(text, font=font)), 1)
        try:
            asc, desc = font.getmetrics()
            th = max(asc + desc, 1)
        except Exception:
            th = max(size, 1)

        img  = Image.new("RGBA", (tw, th), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        try:
            bbox = font.getbbox(text)
            yo   = (th - (bbox[3] - bbox[1])) // 2
        except Exception:
            yo = 0
        draw.text((0, yo), text, font=font, fill=self._hex_to_rgb(color))

        sw = max(int(tw / self.scale_x), 1)
        sh = max(int(th / self.scale_y), 1)
        return ImageTk.PhotoImage(img.resize((sw, sh), Image.LANCZOS))

    def _draw_placeholder(self, field: str, x: float, y: float):
        if field in self.placeholders:
            self.canvas.delete(self.placeholders[field]["item"])
        photo = self._render_placeholder_image(field)
        self._ph_images[field] = photo
        item  = self.canvas.create_image(x, y, image=photo, anchor="center")
        self.canvas.tag_bind(item, "<Button-1>",  lambda e, i=item: self._drag_start(e, i))
        self.canvas.tag_bind(item, "<B1-Motion>", lambda e, i=item: self._drag_move(e, i))
        self.placeholders[field] = {"item": item, "x": x, "y": y}

    def create_placeholder(self, field: str, x=None, y=None, **_):
        if field not in self.fields:
            return
        if x is None or y is None:
            cw  = self.canvas.winfo_width() or 800
            idx = self.fields.index(field)
            x, y = cw // 2, 50 + idx * 60
        self._draw_placeholder(field, x, y)

    def update_preview(self, field: str):
        if field in self.placeholders:
            p = self.placeholders[field]
            self._draw_placeholder(field, p["x"], p["y"])
            self._status(f"Updated field: {field}")

    def _drag_start(self, event, item):
        self._drag = {"item": item, "x": event.x, "y": event.y}

    def _drag_move(self, event, item):
        dx = event.x - self._drag["x"]
        dy = event.y - self._drag["y"]
        self.canvas.move(item, dx, dy)
        self._drag["x"] = event.x
        self._drag["y"] = event.y
        for f, d in self.placeholders.items():
            if d["item"] == item:
                d["x"] += dx
                d["y"] += dy
                break

    # ------------------------------------------------------------------
    # Coordinate scaling
    # ------------------------------------------------------------------
    def _get_scaled_positions(self) -> dict:
        return {
            f: (d["x"] * self.scale_x, d["y"] * self.scale_y)
            for f, d in self.placeholders.items()
        }

    # ------------------------------------------------------------------
    # Draw text
    # ------------------------------------------------------------------
    def _draw_text_on_image(self, img: Image.Image,
                            student: dict, positions: dict) -> Image.Image:
        draw = ImageDraw.Draw(img)
        for field in self.fields:
            if not self.field_vars.get(field, tk.BooleanVar(value=False)).get():
                continue
            if field not in positions:
                continue
            try:
                x, y  = positions[field]
                s     = self.font_settings[field]
                size  = s["size"].get()
                font  = self._get_font(s["font_name"].get(), size)
                text  = student.get(field, "")
                tw    = draw.textlength(text, font=font)
                try:
                    bbox = font.getbbox(text)
                    th   = bbox[3] - bbox[1]
                    yo   = (size - th) // 2
                except Exception:
                    yo = 0
                draw.text((x - tw/2, y - size/2 + yo),
                          text, font=font,
                          fill=self._hex_to_rgb(s["color"].get()))
            except Exception as exc:
                print(f"[draw_text] {field}: {exc}")
        return img

    # ------------------------------------------------------------------
    # Preview
    # ------------------------------------------------------------------
    def preview_certificate(self):
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Load a template first.")
            return
        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Load student data first.")
            return

        img = self._draw_text_on_image(
            self.original_image.copy(), self.excel_data[0],
            self._get_scaled_positions())

        win = tk.Toplevel(self.root)
        win.title("Preview")
        win.transient(self.root)
        win.grab_set()
        win.resizable(True, True)
        win.minsize(400, 300)
        win.configure(bg=C["bg"])

        pw = max(self.root.winfo_width() - 100, 600)
        ph = max(self.root.winfo_height() - 100, 400)
        win.geometry(f"{pw}x{ph}")

        iw, ih   = img.size
        ratio    = min(pw/iw, ph/ih)
        photo    = ImageTk.PhotoImage(
            img.resize((int(iw*ratio), int(ih*ratio)), Image.LANCZOS))

        lbl = tk.Label(win, image=photo, bg=C["bg"], bd=0)
        lbl.image = photo
        lbl.pack(fill="both", expand=True, padx=12, pady=12)

        _btn(win, "Close", win.destroy,
             C["accent"], C["accent2"],
             font_size=9, bold=True,
             padx=24, pady=6).pack(pady=(0, 12))

        win.bind("<Escape>", lambda e: win.destroy())
        win.lift()
        win.focus_set()

    # ------------------------------------------------------------------
    # Generation
    # ------------------------------------------------------------------
    def generate_certificates(self):
        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Load student data first.")
            return
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Load a template first.")
            return
        if not self._gen_lock.acquire(blocking=False):
            messagebox.showwarning("CertWizard", "Generation already in progress.")
            return

        use_cmyk = messagebox.askyesno(
            "Color mode",
            "Generate in CMYK color space?\n\nYes  =  CMYK    No  =  RGB")
        self.color_space.set("CMYK" if use_cmyk else "RGB")

        out_dir = filedialog.askdirectory(title="Select output folder")
        if not out_dir:
            self._gen_lock.release()
            return

        sub = "CMYK" if use_cmyk else "RGB"
        os.makedirs(os.path.join(out_dir, sub), exist_ok=True)

        positions = self._get_scaled_positions()
        total     = len(self.excel_data)
        iw, ih    = self.original_image.size
        pdf_w     = px_to_mm(iw)
        pdf_h     = px_to_mm(ih)

        def _run():
            count = 0
            self._log("Starting generation...", clear=True)
            self._log(f"Records: {total}   Mode: {sub}   Output: {out_dir}")
            self._log("-" * 44)

            for idx, student in enumerate(self.excel_data):
                try:
                    img = self._draw_text_on_image(
                        self.original_image.copy().convert("RGB"),
                        student, positions)

                    pdf = FPDF(unit="mm", format=(pdf_w, pdf_h))
                    pdf.add_page()

                    with tempfile.NamedTemporaryFile(
                            suffix=".png", delete=False) as tmp:
                        tmp_path = tmp.name
                    try:
                        img.save(tmp_path, format="PNG", optimize=True)
                        pdf.image(tmp_path, x=0, y=0, w=pdf_w, h=pdf_h)
                    finally:
                        try:
                            os.remove(tmp_path)
                        except OSError:
                            pass

                    parts = [
                        re.sub(r"[^\w\-_. ]", "",
                               str(student.get(self.fields[i], ""))).strip()
                        for i in range(min(2, len(self.fields)))
                    ]
                    safe = "_".join(p for p in parts if p) or f"cert_{idx+1}"
                    out  = os.path.join(out_dir, sub, f"{safe}_certificate.pdf")
                    pdf.output(out)
                    count += 1
                    self._log(f"[{idx+1}/{total}]  {safe}_certificate.pdf")
                except Exception as exc:
                    self._log(f"[error]  cert {idx+1}: {exc}")

                pct = (idx+1) / total * 100
                self.root.after(0, lambda v=pct: self.progress.configure(value=v))

            self._log("-" * 44)
            self._log(f"Done  {count}/{total} certificates saved.")
            self.root.after(0, lambda: messagebox.showinfo(
                "CertWizard", f"{count} certificate(s) generated successfully."))
            self._gen_lock.release()

        threading.Thread(target=_run, daemon=True).start()

    # ------------------------------------------------------------------
    # Project save / load
    # ------------------------------------------------------------------
    def save_project(self):
        if not self.original_image:
            messagebox.showwarning("Warning", "No template loaded.")
            return
        try:
            project = {
                "version":        "2.1",
                "last_modified":  datetime.now().isoformat(),
                "template_path":  self.template_path,
                "excel_path":     self.excel_path,
                "color_space":    self.color_space.get(),
                "positions":      self._get_scaled_positions(),
                "field_settings": {
                    f: {
                        "size":      self.font_settings[f]["size"].get(),
                        "color":     self.font_settings[f]["color"].get(),
                        "visible":   self.field_vars[f].get(),
                        "font_name": self.font_settings[f]["font_name"].get(),
                    } for f in self.fields
                },
            }
            path = filedialog.asksaveasfilename(
                defaultextension=".certwiz",
                filetypes=[("CertWizard Project", "*.certwiz")])
            if path:
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(project, fh, indent=2)
                messagebox.showinfo("Saved", "Project saved.")
        except Exception as exc:
            messagebox.showerror("Error", f"Save failed:\n{exc}")

    def load_project(self):
        path = filedialog.askopenfilename(
            filetypes=[("CertWizard Project", "*.certwiz")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)

            self.canvas.delete("all")
            self.placeholders.clear()
            self._ph_images.clear()

            tpl = data.get("template_path", "")
            if tpl and os.path.exists(tpl):
                self.load_template(tpl)
            elif tpl:
                messagebox.showwarning("Warning",
                                       f"Template not found:\n{tpl}")

            xl = data.get("excel_path", "")
            if xl and os.path.exists(xl):
                self.load_excel(xl)
            elif xl:
                messagebox.showwarning("Warning",
                                       f"Excel file not found:\n{xl}")

            self.color_space.set(data.get("color_space", "RGB"))

            fs = data.get("field_settings", {})
            for f in self.fields:
                if f in fs:
                    self.font_settings[f]["size"].set(fs[f].get("size", 32))
                    self.font_settings[f]["color"].set(fs[f].get("color", "#000000"))
                    self.font_settings[f]["font_name"].set(
                        fs[f].get("font_name", next(iter(self.available_fonts))))
                    self.field_vars[f].set(fs[f].get("visible", True))

            for field, (sx, sy) in data.get("positions", {}).items():
                if field in self.fields:
                    self._draw_placeholder(
                        field, sx / self.scale_x, sy / self.scale_y)

            messagebox.showinfo("Loaded", "Project loaded.")
        except Exception as exc:
            messagebox.showerror("Error", f"Load failed:\n{exc}")

    # ------------------------------------------------------------------
    # Status / log
    # ------------------------------------------------------------------
    def _status(self, msg: str):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _log(self, msg: str, clear: bool = False):
        def _inner():
            self.info_text.configure(state="normal")
            if clear:
                self.info_text.delete("1.0", tk.END)
            self.info_text.insert(tk.END, msg + "\n")
            self.info_text.see(tk.END)
            self.info_text.configure(state="disabled")
        self.root.after(0, _inner)

    # legacy aliases
    _update_status = _status
    update_status  = _status
    update_info    = lambda self, m, clear=False: self._log(m, clear)
    get_placeholder_positions = _get_scaled_positions


if __name__ == "__main__":
    root = tk.Tk()
    CertificateApp(root)
    root.mainloop()
