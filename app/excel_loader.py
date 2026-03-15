"""
Excel ingestion — isolated from tkinter and PIL.
Returns plain Python dicts; caller owns the UI feedback.
"""
from datetime import datetime

from openpyxl import load_workbook


def read(file_path: str) -> tuple[list, list]:
    """
    Read an xlsx file and return (header, rows).
    header : list of lowercase column names
    rows   : list of dicts  {col_name: str_value}
    Raises ValueError on bad input, propagates openpyxl exceptions.
    """
    wb    = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    header = [
        str(c.value).strip().lower()
        for c in next(sheet.iter_rows(min_row=1, max_row=1))
        if c.value is not None
    ]
    if not header:
        wb.close()
        raise ValueError("No column headers found in the Excel file.")

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rec = {}
        for i, val in enumerate(row):
            if i >= len(header):
                break
            if isinstance(val, datetime):
                val = val.strftime("%d-%m-%Y")
            rec[header[i]] = str(val) if val is not None else ""
        if any(rec.values()):
            rows.append(rec)

    wb.close()

    if not rows:
        raise ValueError("No data rows found in the Excel file.")

    return header, rows
